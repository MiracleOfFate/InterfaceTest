import json
import os
import unittest
import jsonpath
import requests
from openpyxl import load_workbook

from httpop.httpop import httpop
from excel.excelop import excelop


class MyTestCase1(unittest.TestCase):
    listchecklist = []

    # 解析xlsx文件
    @classmethod
    def setUpClass(cls) -> None:
        # 切换到当前路径下
        parentdir = os.path.dirname(__file__)
        parentdir+r"/config/case.xlsx"

        MyTestCase1.listchecklist=excelop().excelread()

        # workbook = load_workbook(r"./config/case.xlsx")
        # sheets = workbook["Sheet1"]
        # rows_sheet = sheets.iter_rows()  # 读取每一行
        #
        # for item in rows_sheet:
        #     if item[0].value == "url":
        #         continue
        #     listes = []
        #     for col in item:
        #         listes.append(col.value)
        #
        #     MyTestCase1.listchecklist.append(listes)

    def test_checklist1(self):
        # print(MyTestCase1.listchecklist)
        # print(os.path.abspath("."))       # 当前绝对路径
        # print(os.path.dirname(__file__))  # 切换到当前路径下
        for checklist in MyTestCase1.listchecklist:
            url = checklist[0]
            expect = checklist[2]
            jsondir = checklist[3]
            # res = requests.get(url)
            res = httpop().api_get(url)
            resjon = json.loads(res.content)
            result = jsonpath.jsonpath(resjon, jsondir)
            # print(resjon)
            # print(type(expect), result[0])
            self.assertEqual(str(expect), result[0])  # 转换成字符串再进行比较


if __name__ == "__main__":
    unittest.main()
