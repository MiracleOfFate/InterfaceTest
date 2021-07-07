from openpyxl import load_workbook


class excelop:
    def excelread(self):
        listchecklist = []
        workbook = load_workbook(r"../config/case.xlsx")
        sheets = workbook["Sheet1"]
        rows_sheet = sheets.iter_rows()  # 读取每一行

        for item in rows_sheet:
            if item[0].value == "url":
                continue
            listes = []
            for col in item:
                listes.append(col.value)

            listchecklist.append(listes)

        return listchecklist